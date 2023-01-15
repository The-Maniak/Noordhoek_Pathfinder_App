import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Specifying the script location:
current_directory = r'C:\Users\wojte\Desktop\python_projects\Noordhoek_Pathfinder\NHP_apps\nl_ports_crewlist_project'
# Determining the source files directories:
arrival_crewlist_directory = os.path.join(current_directory, 'vessel-format-crewlist\\arrival\\')
departure_crewlist_directory = os.path.join(current_directory, 'vessel-format-crewlist\\departure\\')
# Determining the output directory:
output_directory = os.path.join(current_directory, 'ready-document\\')

# Travel document dictionary:
travel_documents = {"PP": "Passport"}

# Dictionary with genders:
genders = {'M': "Male", "F": 'Female'}

# Dictionary with ranks:
ranks = {'Captain': 'Captain', 'C/O SDPO': 'Chief Officer', '2/O SDPO': 'Second Officer',
         '2/O DPO': 'Second Officer', 'C/E': 'Chief Engineer',
         '2/E': 'Second Assistant Engineer, Second Engineer',
         '3/E': 'Third Assistant Engineer, Third Engineer', 'MM': 'Motorman', 'Cook':
             'Cook', 'Bosun': 'Bosun', 'AB Crane': 'Crane Operator', 'AB': 'Able Seaman', }

three_to_two_letters_nationalities = {}


def get_nationalities_codes():
    country_codes_filename = 'country_codes.xlsx'
    country_codes_filepath = os.path.join(current_directory, country_codes_filename)
    wb = load_workbook(country_codes_filepath, data_only=True)
    country_codes_sheet = wb.active
    country_codes_cell = country_codes_sheet["A2":"C250"]
    for nationality, two_letters, three_letters in country_codes_cell:
        three_to_two_letters_nationalities[f'{three_letters.value}'] = f'{two_letters.value}'
    return three_to_two_letters_nationalities


class Crewmember:
    """Class handling the data of all cremembers from both crewlists."""
    dict_all_crew = {}  # Dictionary with all crew
    arrival_crew = {}  # Dictionary with crew arriving to port with the vessel
    departure_crew = {}  # Dictionary with crew departing the port with the vessel

    def __init__(self, surname, name, rank, nationality, date_of_birth, place_of_birth, gender, document_type,
                 document_number, document_state, document_expiration_date, ):
        """This method initializes every person in the crewlist as a separate Crewmember class instance."""
        self.surname = surname
        self.name = name
        self.full_name = f'{surname}, {name}'
        self.rank = ranks.get(rank, 'Other')
        self.nationality = three_to_two_letters_nationalities[nationality]
        self.date_of_birth = date_of_birth
        self.place_of_birth = place_of_birth
        self.gender = genders.get(gender, "Not known")
        self.document_type = travel_documents.get(document_type, 'Passport')
        self.document_number = document_number
        self.document_state = three_to_two_letters_nationalities[document_state]
        self.document_expiration_date = document_expiration_date
        Crewmember.dict_all_crew[self.full_name] = self

    def __str__(self):
        # this was mainly use in development to ensure instances generate as expected.
        return f'{self.surname} {self.name}, {self.rank}'

    @classmethod
    def list_all_crew(cls):
        """Class method used to create an iterable object for filling in the template."""
        return list(Crewmember.dict_all_crew.values())


def read_data_from_vessel_format_crew_list(directory, is_this_arrival):
    """Function reading data from given vessel crew list.
    As the first argument it takes the file directory
    As the second argument it takes the information if it is an arrival or departure crew list."""
    try:
        crew_list_file = os.path.join(directory, os.listdir(directory)[0])
        wb = load_workbook(crew_list_file, data_only=True)
        crew_list_sheet = wb['Crewlist']
        number_of_pob = crew_list_sheet["E61"].value
        port_docs_sheet = wb['Port Doc Copy Sheet']
        cell = port_docs_sheet["B3":f"M{number_of_pob + 2}"]
        for number, surname, name, rank, nationality, date_of_birth, place_of_birth, gender, \
            document_type, document_number, document_state, document_expiration_date in cell:
            person = Crewmember(surname.value, name.value, rank.value, nationality.value, date_of_birth.value,
                                place_of_birth.value,
                                gender.value, document_type.value, document_number.value, document_state.value,
                                document_expiration_date.value, )
            if is_this_arrival is True:
                Crewmember.arrival_crew[person.full_name] = person
            else:
                Crewmember.departure_crew[person.full_name] = person
    except IndexError:
        # No files in the directory
        return


def embarks(person):
    """Function determining how should the 'Embarks at port of call' for each of the crew members.
        It's 'Yes' for persons joining the vessel, 'No' for persons remaining on board or signing off."""
    if person.full_name in Crewmember.arrival_crew.keys():
        return 'No'
    return 'Yes'


def disembarks(person):
    """Function determining how should the 'Disembarks at port of call' be filled for each of the crew members.
        It's 'Yes' for persons signing off, 'No' for persons remaining on board or joining."""
    if person.full_name in Crewmember.departure_crew.keys():
        return 'No'
    return 'Yes'


def write_data_to_template():
    template_directory = r'C:\Users\wojte\Desktop\python_projects\Noordhoek_Pathfinder\NHP_apps\nl_ports_crewlist_project\template'
    template_filename = r'PAX-list PCS-NLl.xlsx'
    image_filepath = os.path.join(template_directory, r'portbase logo.png')
    port_base_logo = Image(image_filepath)
    template_filepath = os.path.join(template_directory, template_filename)
    wb = load_workbook(template_filepath, data_only=True)
    crewlist_sheet = wb['Crew']
    # crewlist_sheet.add_image[port_base_logo, 'V1']
    mapping = {
        'A': 'surname',
        'B': 'name',
        'C': 'gender',
        'D': 'nationality',
        'E': 'date_of_birth',
        'F': 'place_of_birth',
        'G': 'document_type',
        'H': 'document_number',
        'I': 'document_expiration_date',
        'J': 'document_state',
        'S': embarks,
        'T': disembarks,
        'V': 'rank',
    }

    for i, person in enumerate(Crewmember.list_all_crew()):
        for col, attr in mapping.items():
            cell = crewlist_sheet[f'{col}{5 + i}']
            if callable(attr):
                cell.value = attr(person)
            else:
                cell.value = getattr(person, attr)

    wb.save(f'{output_directory}1.xlsx')
    # wb.save(f'{output_directory}1.xls')


def main():
    # Generating the codes with nationalities and document origin countries:
    get_nationalities_codes()
    # Reading data from the arrival crew list:
    read_data_from_vessel_format_crew_list(arrival_crewlist_directory, True)
    # Reading data from the departure crew list:
    read_data_from_vessel_format_crew_list(departure_crewlist_directory, False)
    # Writing data to the NL port template:
    write_data_to_template()


if __name__ == "__main__":
    main()
