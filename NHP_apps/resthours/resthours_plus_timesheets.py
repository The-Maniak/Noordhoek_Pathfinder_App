import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Preparing a dictionary with months for filling in the documents:
months_names = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: "June", 7: 'July', 8: 'August',
                9: 'September', 10: 'October', 11: 'November', 12: 'December'}
months_lengths = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}

# Predefining the catalog with the whole application:
app_directory = r'C:\Users\wojte\Desktop\python_projects\Noordhoek_Pathfinder\NHP_apps\resthours'

# File with data of crew members:
crewlist_filename = 'crewmembers.xlsx'
crewlist_filepath = os.path.join(app_directory, 'crew-data', crewlist_filename)

# File with blank resthours sheet:
resthours_blank_filename = 'NHP_resthours_blank.xlsx'
resthours_blank_filepath = os.path.join(app_directory, "base-files", resthours_blank_filename)

# File with blank timesheets sheets and company logos:
redwise_blank_filename = 'Redwise_blank.xlsx'
redwise_blank_filepath = os.path.join(app_directory, "base-files", redwise_blank_filename)
redwise_logo_filename = 'Redwise_logo.png'
redwise_logo_filepath = os.path.join(app_directory, "base-files", redwise_logo_filename)

scanmar_blank_filename = 'Scanmar_blank.xlsx'
scanmar_blank_filepath = os.path.join(app_directory, "base-files", scanmar_blank_filename)
scanmar_logo_filename = 'Scanmar_logo.png'
scanmar_logo_filepath = os.path.join(app_directory, "base-files", scanmar_logo_filename)

TOS_blank_filename = 'TOS_blank.xlsx'
TOS_blank_filepath = os.path.join(app_directory, "base-files", TOS_blank_filename)
TOS_logo_filename = 'TOS_logo.png'
TOS_logo_filepath = os.path.join(app_directory, "base-files", TOS_logo_filename)

# output folders:
output_folder = os.path.join(app_directory, 'output\\resthours\\')
redwise_output_folder = os.path.join(app_directory, 'output\\Redwise\\')
scanmar_output_folder = os.path.join(app_directory, 'output\\Scanmar\\')
TOS_output_folder = os.path.join(app_directory, 'output\\TOS\\')


# Preparing a list of columns for future iteration:
def excell_column_names(*given_letters):
    """ Function creating a list of MS Excel column names exceeding the single alphabet letters."""
    list_of_columns = []
    alphabet = [chr(x) for x in range(65, 91)]
    for given_letter in given_letters:
        for letter in alphabet:
            list_of_columns.append(f'{given_letter}{letter}')
    return list_of_columns


columns = excell_column_names('', "A", "B")


# Defining functions for filling in resthours (each hour is two cells thus '*2' in the formula):
def filling_resthours(shift_start_time, shift_end_time, sign_on_date, sign_off_date, sign_on_time, sign_off_time,
                      sign_on_time_float, sign_off_time_float):
    # Opening the template and activatin the sheet:
    wb_obj = load_workbook(resthours_blank_filepath)
    sheet_obj = wb_obj.active
    # Inserting the data into cells irrelevant of watch keeping schedule of each seafarer.
    name_cell = sheet_obj['BF6']
    name_cell.value = full_name
    captain_name_cell = sheet_obj['BH25']
    captain_name_cell.value = captain_onboard_cell
    rank_cell = sheet_obj['BM6']
    rank_cell.value = rank.value
    month_and_year = sheet_obj['BF8']
    month_and_year.value = f'{months_names[current_month]} {current_year}'
    # Inserting ex's where the crew member was resting. Leaving blank where he was home.
    if shift_end_time.value < shift_start_time.value:
        "Function filling in a central block of rest time for people working during night shifts"
        for row in sheet_obj.iter_rows(
                min_row=5,
                min_col=(shift_end_time.value + 1) * 2,
                max_col=shift_start_time.value * 2 + 1,
                max_row=number_of_days_in_the_month + 4):
            for rest_cell in row:
                rest_cell.value = 'x'
    else:
        "Function filling in two blocks of rest time for people working during day shift"
        for row in sheet_obj.iter_rows(
                min_row=5,
                min_col=2,
                max_col=shift_start_time.value * 2 + 1,
                max_row=number_of_days_in_the_month + 4):
            for rest_cell in row:
                rest_cell.value = 'x'
        for row in sheet_obj.iter_rows(
                min_row=5,
                min_col=(shift_end_time.value + 1) * 2,
                max_col=49,
                max_row=number_of_days_in_the_month + 4):
            for rest_cell in row:
                rest_cell.value = 'x'
    """"Function clearing the days prior joining the vessel."""
    if sign_on_date:
        try:
            for row in sheet_obj.iter_rows(
                    min_row=5,
                    min_col=2,
                    max_col=49,
                    max_row=sign_on_date.value + 3):
                for at_home_cell in row:
                    at_home_cell.value = ''
            for column in range(2, int(sign_on_time_float.value * 48 + 2)):
                extra_rest_on_sign_on_day = sheet_obj.cell(sign_on_date.value + 4, column)
                extra_rest_on_sign_on_day.value = "x"
            remark_cell = sheet_obj[f"AY{sign_on_date.value + 4}"]
            remark_cell.value = f"Signed on at {sign_on_time[0:5]}"
        except TypeError as e:
            # print(f'At {full_name} in the sign on section, there was an error: {e}')
            pass
        """"Function clearing the days after leaving the vessel."""
    if sign_off_date is not None:
        try:
            for row in sheet_obj.iter_rows(
                    min_row=sign_off_date.value + 5,
                    min_col=2,
                    max_col=49,
                    max_row=number_of_days_in_the_month + 4):
                for at_home_cell in row:
                    at_home_cell.value = ''
            for column in range(int(sign_off_time_float.value * 48 + 1), 48):
                extra_rest_on_sign_on_day = sheet_obj.cell(sign_off_date.value + 4, column)
                extra_rest_on_sign_on_day.value = "x"
            remark_cell = sheet_obj[f"AY{sign_off_date.value + 4}"]
            remark_cell.value = f"Signed off at {sign_off_time[0:5]}"
        except TypeError as e:
            # print(f'At {full_name} in the sign off section, there was an error {e}.')
            pass
    wb_obj.save(f'{output_folder}{full_name}.xlsx')


# Filling Redwise and Scanamar timesheets.
def filling_redwise_scanmar(template_filepath, image_filepath, image_position, output_folder, shift_start_time,
                            shift_end_time):
    """Function filling Redwise & Scanmar timesheets, one function for two as they have quite similar layout.
    As there are different templates and images, they are loaded outside of the function, in the very last for loop
    of the code. This function fills in the hours in the part, which is same for both templates."""
    wb_obj = load_workbook(template_filepath, data_only=True)
    sheet_obj = wb_obj.active
    logo = Image(image_filepath)
    sheet_obj.add_image(logo, image_position)
    name_cell = sheet_obj['B6']
    name_cell.value = full_name
    rank_cell = sheet_obj['B7']
    rank_cell.value = rank.value
    month = sheet_obj['B5']
    month.value = f'{months_names[current_month]}'
    year = sheet_obj['G5']
    year.value = f'{current_year}'
    formula = sheet_obj['C42']
    formula.value = '=SUM(C11:C41)'
    if sign_on_date.value is None:
        start_filling_row = 11
    else:
        start_filling_row = sign_on_date.value + 11
        working_hours_sign_on_day = sheet_obj[f"B{start_filling_row - 1}"]
        amount_hours_sign_on_day = sheet_obj[f"C{start_filling_row - 1}"]
        comment_sign_on_day = sheet_obj[f"E{start_filling_row - 1}"]
        comment_sign_on_day.value = f'Signed on at {sign_on_time[0:5]}'
        if shift_end_time < shift_start_time:
            if sign_on_time_float.value * 24 < shift_start_time:
                working_hours_sign_on_day.value = f'{shift_start_time} - {shift_end_time}'
                amount_hours_sign_on_day.value = abs(shift_end_time - shift_start_time)
            else:
                working_hours_sign_on_day.value = ''
                amount_hours_sign_on_day.value = ''
        else:
            if sign_on_time_float.value * 24 < shift_start_time:
                working_hours_sign_on_day.value = f'{shift_start_time} - {shift_end_time}'
                amount_hours_sign_on_day.value = abs(shift_end_time - shift_start_time)
            elif sign_on_time_float.value * 24 < shift_end_time:
                working_hours_sign_on_day.value = f'{sign_on_time[0:5]} - {shift_end_time}:00'
                amount_hours_sign_on_day.value = abs(sign_on_time_float.value * 24 - shift_end_time)
            else:
                working_hours_sign_on_day.value = ''
                amount_hours_sign_on_day.value = ''
    if sign_off_date.value is None:
        end_filling_row = number_of_days_in_the_month + 10
    else:
        end_filling_row = sign_off_date.value + 9
        working_hours_sign_off_day = sheet_obj[f'B{end_filling_row + 1}']
        amount_hours_sign_off_day = sheet_obj[f'C{end_filling_row + 1}']
        comment_sign_on_day = sheet_obj[f'E{end_filling_row + 1}']
        comment_sign_on_day.value = f'Signed off at {sign_off_time[0:5]}'
        if shift_start_time < shift_end_time:
            if sign_off_time_float.value * 24 < shift_start_time:
                working_hours_sign_off_day.value = ''
                amount_hours_sign_off_day.value = ''
            elif sign_off_time_float.value * 24 < shift_end_time:
                working_hours_sign_off_day.value = f'{shift_start_time}:00 - {sign_off_time[0:5]}'
                amount_hours_sign_off_day.value = f'{sign_off_time_float.value * 24 - shift_start_time}'
            else:
                working_hours_sign_off_day.value = f"{shift_start_time}:00 - {shift_end_time}:00"
                amount_hours_sign_off_day.value = abs(shift_end_time - shift_start_time)
        else:
            working_hours_sign_off_day.value = f'{shift_start_time}:00 - {shift_end_time}:00'
            amount_hours_sign_off_day.value = abs(shift_end_time - shift_start_time)
    for row in sheet_obj.iter_rows(min_row=start_filling_row, min_col=2, max_row=end_filling_row, max_col=2):
        for cell in row:
            cell.value = f'{shift_start_time}:00 - {shift_end_time}:00'
    for row in sheet_obj.iter_rows(min_row=start_filling_row, min_col=3, max_row=end_filling_row, max_col=3):
        for cell in row:
            cell.value = abs(shift_end_time - shift_start_time)
    wb_obj.save(f'{output_folder}{full_name}.xlsx')


# Filling TOS agency timesheet.
def filling_TOS(shift_start_time, shift_end_time):
    """Function filling TOS timesheet."""
    wb_obj = load_workbook(TOS_blank_filepath, data_only=True)
    sheet_obj = wb_obj.active
    name_cell = sheet_obj['E9']
    name_cell.value = full_name
    month_and_year = sheet_obj['E11']
    month_and_year.value = f'{months_names[current_month]} {current_year}'
    TOS_logo = Image(TOS_logo_filepath)
    sheet_obj.add_image(TOS_logo, 'A1')
    if sign_on_date.value is None:
        start_filling_row = 15
    else:
        start_filling_row = sign_on_date.value + 15
        sign_on_date_start_cell = sheet_obj[f'B{start_filling_row - 1}']
        sign_on_date_start_cell.value = f'{sign_on_time[0:5]}' if sign_on_time_float.value * 24 > shift_start_time \
            else f'{shift_start_time}'
        sign_on_date_end_cell = sheet_obj[f'C{start_filling_row - 1}']
        sign_on_date_end_cell.value = f'{shift_end_time}:00' if sign_on_time_float.value * 24 > shift_start_time \
            else f''
        sign_on_date_hour_cell = sheet_obj[f'D{start_filling_row - 1}']
        sign_on_date_hour_cell.value = '' if sign_on_time_float.value * 24 > shift_end_time \
            else f'{shift_end_time - sign_on_time_float.value * 24}'.rstrip('0').rstrip('.')
        sign_on_date_comment_cell = sheet_obj[f'I{start_filling_row - 1}']
        sign_on_date_comment_cell.value = f'Signed on at {sign_on_time[0:5]}'
    if sign_off_date.value is None:
        end_filling_row = number_of_days_in_the_month + 14
    else:
        end_filling_row = sign_off_date.value + 13
        sign_off_date_start_cell = sheet_obj[f'B{end_filling_row + 1}']
        sign_off_date_start_cell.value = f'{shift_start_time}:00' if sign_off_time_float.value * 24 > shift_start_time \
            else shift_start_time
        sign_off_date_end_cell = sheet_obj[f'C{end_filling_row + 1}']
        sign_off_date_end_cell.value = f'{sign_off_time[0:5]}' if sign_off_time_float.value * 24 < shift_end_time \
            else shift_end_time
        sign_off_date_hours_cell = sheet_obj[f'D{end_filling_row + 1}']
        if sign_off_time_float.value * 24 < shift_start_time:
            sign_off_date_hours_cell.value = ''
        elif shift_start_time < sign_off_time_float.value * 24 < shift_end_time:
            sign_off_date_hours_cell.value = f'{sign_off_time_float.value * 24 - shift_start_time}'.rstrip('0').rstrip(
                '.')
        else:
            sign_off_date_hours_cell.value = f'{abs(shift_end_time - shift_start_time)}'
        sign_off_date_comment_cell = sheet_obj[f'I{end_filling_row + 1}']
        sign_off_date_comment_cell.value = f'Signed off at {sign_off_time[0:5]}'
    for row in sheet_obj.iter_rows(min_row=start_filling_row, min_col=2, max_row=end_filling_row, max_col=2):
        for cell in row:
            cell.value = f'{shift_start_time}:00'
    for row in sheet_obj.iter_rows(min_row=start_filling_row, min_col=3, max_row=end_filling_row, max_col=3):
        for cell in row:
            cell.value = f'{shift_end_time}:00'
    for row in sheet_obj.iter_rows(min_row=start_filling_row, min_col=4, max_row=end_filling_row, max_col=4):
        for cell in row:
            cell.value = abs(shift_end_time - shift_start_time)
    wb_obj.save(f'{TOS_output_folder}{full_name}.xlsx')


"""Obtaining data of crewmembers from the source file."""
print('Okay, rest hours and agency time sheets - lets get started...')
wb = load_workbook(crewlist_filepath, data_only=True)
sheet = wb.active
number_of_pob = sheet["B1"].value
captain_onboard_cell = sheet['D1'].value
current_month = sheet['F1'].value
number_of_days_in_the_month = months_lengths[current_month]
current_year = sheet['I1'].value
cell = sheet["A3":f"L{number_of_pob + 2}"]
# Reading the data in source file and filling in individual timesheets:
for surname, name, rank, shift_start_time, shift_end_time, sign_on_date, sign_on_time, sign_on_time_float, \
    sign_off_date, sign_off_time, sign_off_time_float, agency in cell:
    full_name = f"{surname.value}, {name.value}"
    sign_on_time = str(sign_on_time.value)
    sign_off_time = str(sign_off_time.value)
    filling_resthours(shift_start_time, shift_end_time, sign_on_date, sign_off_date, sign_on_time, sign_off_time,
                      sign_on_time_float, sign_off_time_float)
    """"This part of the script is responsible bo filling in timesheets for crewmembers."""
    # Redwise
    if agency.value == "Redwise":
        filling_redwise_scanmar(redwise_blank_filepath, redwise_logo_filepath, 'I1', redwise_output_folder,
                                shift_start_time.value, shift_end_time.value)
    # Scmanmar
    if agency.value == 'Scanmar':
        filling_redwise_scanmar(scanmar_blank_filepath, scanmar_logo_filepath, 'J1', scanmar_output_folder,
                                shift_start_time.value, shift_end_time.value)
    # TOS
    if agency.value == "TOS":
        filling_TOS(shift_start_time.value, shift_end_time.value)
print('Sheets ready.')
